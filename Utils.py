import openpyxl
from time import clock
import pandas as pd


class Utils:
    def __init__(self, prefix):
        self.path = prefix[0]
        self.ext_xlsx = ".xlsx"
        self.result = prefix[1]

    """
    get_excel : read excel file by path
    :param 
        path
    :return
        panda object
    """
    def get_excel(self):
        return pd.read_excel(self.path)
    
    # TODO
    def get_excel2(self):
        excel_file = openpyxl.load_workbook(self.path)
        sheet_names = excel_file.get_sheet_names()
        print(str(sheet_names))

    """
    get_excel_to_dict : read excel file then make data to dict
    :param
        path
    :return
        dict
    """
    def get_excel_to_dict(self):
        # print(pd.read_excel(self.path).to_dict())
        return pd.read_excel(self.path).to_dict()

    """
    make_excel : make dictionary data to excel file
    :param
         path : 
         ext : .xlsx
         result :
    :return
        excel file in path
    """
    def make_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="INDEX")
        sheet.cell(row=row, column=2, value="barcode")
        sheet.cell(row=row, column=3, value="guide")
        sheet.cell(row=row, column=4, value="EA")
        sheet.cell(row=row, column=5, value="Library barcode")
        sheet.cell(row=row, column=6, value="Library guide")
        sheet.cell(row=row, column=7, value="barcode_ox")
        sheet.cell(row=row, column=8, value="Guide_ox")
        sheet.cell(row=row, column=9, value="guide_index")
        """
        key : ['index', 'barcode', guide , ea, 'lib_barcode', 'lib_quide', 'O', 'X', []]
        , 0 : ['A_GAAGTA_PAM_7', 'ATGTCATACATACTC', 'TGAGTGGGCTTAGGAGGGG' , 1, 'ATGTCATACATACTC', 'GGAGGGAGCTGGGTTTTAG', 'O', 'X', []]
        """
        row = 2
        for key, val_arr in sorted(self.result.items()):
            sheet.cell(row=row, column=1, value=val_arr[0])
            sheet.cell(row=row, column=2, value=val_arr[1])
            sheet.cell(row=row, column=3, value=val_arr[2])
            sheet.cell(row=row, column=4, value=val_arr[3])
            sheet.cell(row=row, column=5, value=val_arr[4])
            sheet.cell(row=row, column=6, value=val_arr[5])
            sheet.cell(row=row, column=7, value=val_arr[6])
            sheet.cell(row=row, column=8, value=val_arr[7])
            col = 8
            for g_idx in val_arr[8]:
                if isinstance(g_idx, list):
                    for idx in g_idx:
                        col = col + 1
                        sheet.cell(row=row, column=col, value=str(idx))
                else:
                    col = col + 1
                    sheet.cell(row=row, column=col, value=str(g_idx))
            row = row + 1

        workbook.save(filename=self.path + str(clock()) + self.ext_xlsx)


